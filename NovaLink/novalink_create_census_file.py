import openpyxl
from openpyxl.styles import Font
import os
import pymysql
from datetime import datetime
from datetime import time

# Database connection parameters
DB_HOST = "birdprod.csuausqb2ywx.us-east-1.rds.amazonaws.com"
DB_PORT = 3306
DB_USER = "ladybird"
DB_PASSWORD = "SunsetInTurk3y!"
DB_SCHEMA = "birdprod"

# Census template path
CENSUS_TEMPLATE_PATH = "C:/Automation/NovaLink/template_files/2023_PROCENSUS_TEMPLATE.xlsx"

# Census output paths
CENSUS_OUTPUT_PATH_TEST = "C:/Automation/NovaLink/test_census/"
CENSUS_OUTPUT_PATH_FULL = "C:/Automation/NovaLink/full_census/"

def append_from_excel():
    try:
        # Connect to the database
        conn = pymysql.connect(
            host=DB_HOST,
            port=DB_PORT,
            user=DB_USER,
            password=DB_PASSWORD,
            database=DB_SCHEMA
        )
        cursor = conn.cursor()

        # Query to fetch data from the database and get census_mode and invitation_cd
        query = """
            SELECT 
                id as invitation_id, 
                company_id, 
                invitation_cd, 
                census_mode
            FROM 
                invitation_dev_kiandre 
            WHERE 
                (inactive_yn = 0 
                AND individual_step_complete_yn = 1 
                AND paycheck_step_complete_yn = 1 
                AND auth_complete_yn = 1 
                AND ((census_mode = 'TEST' 
                        AND decoder_submitted_dt IS NOT NULL 
                        AND test_census_create_dt IS NULL) 
                    OR (census_mode = 'FULL' 
                        AND decoder_submitted_dt IS NOT NULL 
                        AND full_census_request_dt IS NOT NULL 
                        AND full_census_create_dt IS NULL)))
            LIMIT 1;
        """
        cursor.execute(query)
        result = cursor.fetchone()

        if result:
            invitation_id, company_id, invitation_cd, census_mode = result
        else:
            print("No matching record found in the database.")
            return

        # Generate target Excel file name
        batch_id = datetime.now().strftime("%Y-%m-%d_%H.%M.%S")
        target_file = os.path.join(CENSUS_OUTPUT_PATH_FULL if census_mode == "FULL" else CENSUS_OUTPUT_PATH_TEST,
                                   f"{invitation_cd}_NOVALINK_CENSUS_{batch_id}.xlsx")

        # Load census template
        source_workbook = openpyxl.load_workbook(CENSUS_TEMPLATE_PATH, read_only=True)
        source_sheet = source_workbook.active

        # Create or load target workbook
        if os.path.exists(target_file):
            target_workbook = openpyxl.load_workbook(target_file)
        else:
            target_workbook = openpyxl.Workbook()

        # Get or create target sheet
        target_sheet = target_workbook.active if 'target_sheet' not in locals() else target_workbook.create_sheet(title='Employee')

        # Clear existing data in the target sheet except the title row
        if target_sheet.max_row > 1:
            target_sheet.delete_rows(2,target_sheet.max_row)

        for row in source_sheet.iter_rows(max_row =1, values_only = True):
            target_sheet.append(row)

        for cell in target_sheet[1]:
            cell.font = Font(bold=True)

        # Query to fetch data from the database
        census_query = """
            SELECT 
                First_Name, 
                Last_Name, 
                SSN, 
                Gender, 
                Date_Of_Birth, 
                Original_Date_Of_Hire, 
                Date_of_Separation, 
                Separation_Due_to_Death_Disability_or_Retirement, 
                Date_Of_Rehire, 
                Employee_Involuntarily_Separated, 
                Hours, 
                Gross_Compensation, 
                `W-2_Box_1_Compensation`, 
                `Pre-Entry_Compensation`, 
                Pre_Tax_Health125_Qual_Transportation_132f4, 
                Reimbursements_Allowances_and_Fringe_Benefits, 
                Bonus, 
                Commissions, 
                Overtime_Compensation, 
                Custom_Excluded_Compensation, 
                Severance_Compensation, 
                Pre_Tax_Salary_Deferral, 
                Roth_Salary_Deferral, 
                Employer_Match, 
                Safe_Harbor_Contribution, 
                After_Tax_Contribution, 
                Employer_Contribution, 
                Ownership_Percent, 
                Officer, 
                Excluded_Employee, 
                Job_Classification, 
                Company_Division, 
                Union_Employee, 
                Address_1, 
                Address_2, 
                City, 
                State, 
                Zip_Code, 
                Personal_Email, 
                Work_Email, 
                Phone_Number, 
                company_id, 
                individual_id 
            FROM 
                raw_census_vw  
            WHERE 
                company_id = %s
        """
        cursor.execute(census_query, (company_id,))
        data = cursor.fetchall()

        # Append data from database to target sheet
        for row in data:
            target_sheet.append(row)



        # Save changes to the workbook
        target_workbook.save(target_file)
        print("Data appended successfully.")
    except Exception as e:
        print(f"Error: {e}")
    finally:
        # Close cursor and connection
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()


# Call the function
append_from_excel()



